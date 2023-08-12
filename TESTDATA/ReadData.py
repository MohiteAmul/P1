import  openpyxl

book=openpyxl.load_workbook('C:\\Users\\Amul\\Desktop\\Book1.xlsx')
sheet=book.active
cell=sheet.cell(row=1,column=2)
Dict={}
# print(cell.value)
#
# chnage=cell.value='Amul'
# print(chnage)
# print(sheet.max_row)
# print(sheet.max_column)

#
# for i in range(1,sheet.max_row+1):
#     for j in range(1,sheet.max_column):
#         print(sheet.cell(row=i,column=j).value)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value=="T2":
        for j in range(2,sheet.max_column+1):
            # print(sheet.cell(row=i,column=j).value)
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
print(Dict)