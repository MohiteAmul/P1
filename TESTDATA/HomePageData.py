import openpyxl


class HomePageData:

    @staticmethod
    def getTestdata(test_case_name):
        book = openpyxl.load_workbook('C:\\Users\\Amul\\Desktop\\Book1.xlsx')
        sheet = book.active
        cell = sheet.cell(row=1, column=2)
        Dict = {}
        for i in range(1, sheet.max_row):
            if sheet.cell(row=i, column=1).value == test_case_name:
                # FNAME=sheet.cell(row=1,column=2).value
                # LNAME=sheet.cell(row=i,column=2).value
                # Dict.append({'F.NAME':FNAME,'L.NAME':LNAME})
                # break
                #

                for j in range(2, sheet.max_column + 1):
                    # print(sheet.cell(row=i,column=j).value)
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
